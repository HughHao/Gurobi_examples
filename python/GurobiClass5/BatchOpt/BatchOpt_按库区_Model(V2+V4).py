
# -*- coding: utf-8 -*-
"""
@author: help@gurobi.cn

"""
from gurobipy import *
from collections import OrderedDict
import time
import openpyxl

def ReadData(DataPath, Packages, Warehouse, Areas, Patterns, Record):
    print ("ReadData !")
    data = openpyxl.load_workbook(DataPath)
    table = data["Sheet1"]
    AreasCnt = 0
    WarehouseCnt = 0
    PackagesCnt = 0
    PatternsCnt = 0
    for i in table.iter_rows(min_row=2, max_col=3):
        row =[i[0].value, i[1].value, i[2].value] #抓取行
        Record.append(row)
        if row[1] not in Areas.keys():
            Areas[row[1]] = [AreasCnt, []]
            AreasCnt += 1
        temp = row[1].split('-')    #统计仓库
        if temp[0] not in Warehouse.keys():
            Warehouse[temp[0]] = [WarehouseCnt, []]
            WarehouseCnt += 1
        
        if row[0] in Packages.keys():
            Packages[row[0]][1] += row[2]
            if Warehouse[temp[0]][0] not in Packages[row[0]][2]: 
                Packages[row[0]][2].append(Warehouse[temp[0]][0])
                      
            if Areas[row[1]][0] not in Packages[row[0]][3]:  
                Packages[row[0]][3].append(Areas[row[1]][0])
        else:
            Packages[row[0]] = [PackagesCnt, row[2], [Warehouse[temp[0]][0]], [Areas[row[1]][0]]]
            PackagesCnt += 1
            
    WarehouseSort = list(Warehouse.keys())           
    for i in Packages.keys():
        temp = list(Packages[i][2])
        temp.sort()
        temp.insert(0, Packages[i][1])
        CheckPatterns = tuple(temp)
        if CheckPatterns in Patterns:
            Patterns[CheckPatterns].append(i)
        else:
            Patterns[CheckPatterns] = [PatternsCnt, i]
            for j in range(len(Packages[i][2])):
                Warehouse[WarehouseSort[Packages[i][2][j]]][1].append(CheckPatterns)
            PatternsCnt += 1
            
def BuildFirstStageModel(Packages, Warehouse, Patterns, Batches, PackageLimit, GoodsLimit, BatchePattern, SolveTime):
    print ("BuildModel Version 2!")
    XINDEX = {}                                                               #变量下标
    YINDEX = {}                                                               #变量下标 
    for j in range(1, Batches[1]+1):
        for i in Patterns.keys():
            XINDEX[Patterns[i][0],j] = i[0]     
        for k in range(len(Warehouse)):
            YINDEX[j,k] = 1

    model = Model()
    x = model.addVars(XINDEX.keys(), vtype=GRB.INTEGER, name='x')             #变量x_{pj}
    y = model.addVars(YINDEX.keys(), obj=YINDEX, vtype=GRB.BINARY, name='y')  #变量y_{jk} 
    z = model.addVars(Batches[1], vtype=GRB.BINARY, name='z')                 #变量z_{j}

    #约束(1) 保证分配的波次数量
    if Batches[0] == Batches[1]:
        model.addConstr(z.sum('*') == Batches[0])
    else:
        model.addConstr(z.sum('*') >= Batches[0])
        model.addConstr(z.sum('*') <= Batches[1])

    #约束(2) 保证模式使用次数和其包裹数量匹配
    for i in Patterns.keys():
        model.addConstr(x.sum(Patterns[i][0],'*') == len(Patterns[i])-1)
      
    #约束(3) 保证单一波次商品件数在区间[G1,G2]中
    for j in range(1, Batches[1]+1):
        model.addConstr(x.prod(XINDEX,'*',j) >= z[j-1]*GoodsLimit[0])
        model.addConstr(x.prod(XINDEX,'*',j) <= z[j-1]*GoodsLimit[1])
    
    #约束(4) 保证单一波次包裹数量在区间[P1,P2]中
    for j in range(1, Batches[1]+1):
        model.addConstr(x.sum('*',j) >= z[j-1]*PackageLimit[0])
        model.addConstr(x.sum('*',j) <= z[j-1]*PackageLimit[1])
    
    #约束(5) 确认波次j是否用到仓库k
    WarehouseSort = list(Warehouse.keys())  
    for j in range(1, Batches[1]+1):    
        for k in range(len(Warehouse)):
            expr = LinExpr()
            for p in range(len(Warehouse[WarehouseSort[k]][1])):
                expr += x[Patterns[Warehouse[WarehouseSort[k]][1][p]][0],j]
            model.addConstr(y[j,k] >= 0.001*expr)
    

    model.setParam(GRB.Param.TimeLimit, SolveTime)   #求解时间
    model.optimize()
    
    #获取求解结果
    for j in range(1, Batches[1]+1):
        if z[j-1].x != 0:
            for i in Patterns.keys():
                value = x[Patterns[i][0],j].x
                if value != 0:
                    if j not in BatchePattern.keys():
                        BatchePattern[j] = [[i, round(value)]]
                    else:
                        BatchePattern[j].append([i, round(value)])

def BuildSecondStageModel(Packages, Areas, Patterns, BatchePattern, Solutions, SolveTime):
    print ("BuildModel Version 4!")
    XINDEX = {}
    YINDEX = {}
    AreaXINDEX = {}  
    for j in BatchePattern.keys():
        for i in range(len(BatchePattern[j])):
            for ii in range(1, len(Patterns[BatchePattern[j][i][0]])):
                PackageKey = Patterns[BatchePattern[j][i][0]][ii]
                XINDEX[Packages[PackageKey][0], j] = BatchePattern[j][i][1]          
                for iii in range(len(Packages[PackageKey][3])):
                    if (j, Packages[PackageKey][3][iii]) not in AreaXINDEX.keys():
                        AreaXINDEX[j, Packages[PackageKey][3][iii]] = [(Packages[PackageKey][0], j)]
                    else:
                        AreaXINDEX[j, Packages[PackageKey][3][iii]].append((Packages[PackageKey][0], j))
        for k in range(len(Areas)):
            YINDEX[j,k] = 1
            
    model = Model()
    x = model.addVars(XINDEX.keys(), vtype=GRB.BINARY, name='x')              #变量x_{ij}, 包裹j是否分配到波次j
    y = model.addVars(YINDEX.keys(), obj=YINDEX, vtype=GRB.BINARY, name='y')  #变量y_{jk} 
    
    #约束(1) 保证包裹数量要和模式使用次数匹配
    for j in BatchePattern.keys():
        for i in range(len(BatchePattern[j])):
            expr = LinExpr()
            for ii in range(1, len(Patterns[BatchePattern[j][i][0]])):
                PackageKey = Patterns[BatchePattern[j][i][0]][ii]
                expr += x[Packages[PackageKey][0], j]
            model.addConstr(expr == BatchePattern[j][i][1])
    
    #约束(2) 保证每个包裹都安排到波次中去
    for i in Packages.keys():
        model.addConstr(x.sum(Packages[i][0],'*') == 1)
        
    
    #约束(3) 确认波j是否用到库区k
    for i in AreaXINDEX.keys():
        expr = LinExpr()
        for j in range(len(AreaXINDEX[i])):
            expr += x[AreaXINDEX[i][j]]
        model.addConstr(y[i] >= 0.001*expr)
    
            
    model.setParam(GRB.Param.TimeLimit, SolveTime)
    model.optimize()
     
    PackagesSort = list(Packages.keys())
    for key in XINDEX.keys():
        if x[key].x != 0:
            Solutions[PackagesSort[key[0]]] = key[1]
                   
         
def OutputResult(Solutions, Record):
    workbook = openpyxl.Workbook()
   
    worksheet = workbook.active
    worksheet.title = "方案"
    worksheet.cell(row=1, column=1).value = 'pakage_no'
    worksheet.cell(row=1, column=2).value = 'warehouse'
    worksheet.cell(row=1, column=3).value = 'goods_qty'
    worksheet.cell(row=1, column=4).value = '波次'
    
    for i in range(len(Record)):
        worksheet.cell(row=i+2, column=1).value  = Record[i][0]
        worksheet.cell(row=i+2, column=2).value  = Record[i][1]
        worksheet.cell(row=i+2, column=3).value  = Record[i][2]
        worksheet.cell(row=i+2, column=4).value  = Solutions[Record[i][0]]
    
    workbook.save('Result_按库区(V2+V4).xlsx')
    
       
try:
    DataPath = '测试案例.xlsx'                     #数据
    Packages =  OrderedDict()                     #包裹信息
    Warehouse = OrderedDict()                     #仓库信息
    Areas = OrderedDict()                         #区域信息
    Patterns = OrderedDict()                      #模式
    Record = []                                   #数据记录                            
    BatchePattern = {}                            #第一阶段波次-模式对应关系
    Solutions = {}                                #解
    Batches = [100, 120]                          #波次数量
    PackageLimit = [500, 550]                     #单一波次包裹限制
    GoodsLimit = [1800,3000]                      #单一波次商品件数限制
    FirstStageSolveTime = 100                     #求解时间
    SecondStageSolveTime = 100                    #求解时间
    
    starttime = time.time()
    ReadData(DataPath, Packages, Warehouse, Areas, Patterns, Record)
    BuildFirstStageModel(Packages, Warehouse, Patterns, Batches, PackageLimit, GoodsLimit, BatchePattern, FirstStageSolveTime)
    BuildSecondStageModel(Packages, Areas, Patterns, BatchePattern, Solutions, SecondStageSolveTime)
    OutputResult(Solutions, Record)
    endtime = time.time()
    print ("运行结束! 总耗时 = ", endtime-starttime)
    

except GurobiError as exception:
    print('Error code ' + str(exception.errno) + ": " + str(exception))

except AttributeError:
    print('Encountered an attribute error')
    

